'''
Экспорт данных их ЭЭГ студио настраивается через ЭЭГ студио, это надо будет учесть!!!
'''


import os
import pyautogui
import time
import subprocess

import copy
from PIL import Image as PILImage


from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from openpyxl.drawing.image import Image
from openpyxl.chart import (
    LineChart,
    Reference,
)


import csv

names = []
with open('c:/Пилоты код/poleti.csv', 'r', newline='') as csvfile:
    spamreader = csv.reader(csvfile, delimiter=';', quotechar='|')
    for row in spamreader:
        names.append(row)
num_rows_names = len(names)
num_cols_names = len(names[0])

core_dir = 'C:/Пилоты'
progpath_eeg = "C:/Program Files (x86)/Mitsar/EEGStudio/Modules/Process/EEGStudio.exe"
resources = 'C:/Пилоты_обработано/таблицы' 

screen1_path = resources + "/screen1.png"
screen2_path = resources + "/screen2.png"
screen3_path = resources +"/screen3.png"
screen_timescale_path = resources+"/timescale.png"
ref_size = os.path.getsize('C:/Пилоты_обработано/Обработано/Безрученкова Татьяна/2.1 день/241105_0030_EEG.mdfx')
small_files = []

for surname in os.listdir(core_dir):
    #print(surname)
    day_dir = os.listdir(os.path.join(core_dir, surname).replace('\\','/'))
    for day in day_dir:
        #print(day)
        DIRECTORY_TO_LOOP = os.path.join(core_dir, surname, day).replace('\\','/')

        
        path_to_excel = os.path.join(resources, surname + ' '+ day).replace('\\','/') + '.xlsx' # поменять path, сделать отдельную директорию куда скидывать все таблица
        wb1 = Workbook()
        wb1.save(path_to_excel)
        wb1.close()
        
        for root, dirs, files in os.walk(DIRECTORY_TO_LOOP, topdown=False):
            for file in files:
                if file.endswith('.mdfx'):
                    if os.path.getsize(os.path.normpath(os.path.join(root, file)).replace('\\','/')) < ref_size :
                           print(file)
                    else:
                        filepath_eeg = os.path.normpath(os.path.join(root, file))
                        path = filepath_eeg.split(os.sep)
                        file_name = path[4].split('.')[0]#path поменять
                        for ind in range (num_rows_names):
                            if names[ind][3] == file_name:
                                name_str = names[ind]
                                names.pop(ind)
                                num_rows_names = len(names)
                                break
                                
                        current_day = DIRECTORY_TO_LOOP




                        p = subprocess.Popen([progpath_eeg, filepath_eeg])


                        time.sleep(15)
                        pyautogui.moveTo(196,347,0.5)
                        pyautogui.click(button = 'left')
                        time.sleep(0.5)
                        pyautogui.moveTo(196,602,0.5)
                        pyautogui.click(button = 'left')
                        time.sleep(0.5)
                        pyautogui.moveTo(196,845,0.5)
                        pyautogui.click(button = 'left')
                        time.sleep(0.5)



                        shift = 250
                        time.sleep(5)
                        Xpos1 = 303
                        Ypos1 = 320
                        Xpos1_exp = 411
                        Ypos1_exp = 534
                        pyautogui.moveTo(Xpos1,Ypos1, 0.5)
                        pyautogui.click(button = 'right')
                        pyautogui.moveTo(Xpos1_exp,Ypos1_exp, 0.5)
                        pyautogui.click(button = 'left')
                        time.sleep(3)

                        pyautogui.keyDown('1')
                        pyautogui.keyDown('enter')

                        time.sleep(1)
                        pyautogui.screenshot(screen1_path , region=(1,228, 1812, 240))

                        Xpos1 = 303
                        Ypos1 = 320 +shift
                        Xpos1_exp = 411
                        Ypos1_exp = 785

                        time.sleep(2)
                        pyautogui.moveTo(Xpos1,Ypos1,0.5)
                        pyautogui.click(button = 'right')
                        pyautogui.moveTo(Xpos1_exp,Ypos1_exp,0.5)
                        pyautogui.click(button = 'left')
                        time.sleep(3)

                        pyautogui.keyDown('2')
                        pyautogui.keyDown('enter')

                        time.sleep(1)
                        pyautogui.screenshot(screen2_path , region=(1,224+shift, 1812, 240))

                        Xpos1 = 303
                        Ypos1 = 320+2*shift
                        Xpos1_exp = 411
                        Ypos1_exp = 867

                        time.sleep(2)
                        pyautogui.moveTo(Xpos1,Ypos1,0.5)
                        pyautogui.click(button = 'right')
                        pyautogui.moveTo(Xpos1_exp,Ypos1_exp,0.5)
                        pyautogui.click(button = 'left')


                        pyautogui.keyDown('3')
                        pyautogui.keyDown('enter')
                        time.sleep(1)
                        pyautogui.screenshot(screen3_path , region=(1,224+2*shift, 1812, 240))

                        pyautogui.screenshot(screen_timescale_path , region=(22,964, 1850, 30))
                        



                        time.sleep(0.5)
                        p.terminate()
                        

                        size1= 1812
                        size2 = 1850
                        new_size = (1812-149, 30)
                        im_timescale = PILImage.open(screen_timescale_path)
                        im_timescale.thumbnail(new_size)
                        screen_path = [screen1_path, screen2_path, screen3_path]
                        for iterator in range(3):
                            
                            im_specter = PILImage.open(screen_path[iterator])

                            img_conct = PILImage.new("RGB", (1812, im_specter.size[1]+im_timescale.size[1]))
                            img_conct.paste(im_specter, (0,0))
                            img_conct.paste(im_timescale, (149,im_specter.size[1]))
                            img_conct.save(resources+"/screen" +  str(iterator+1) + ".png")
                        

                        


                        wb = load_workbook(filename = path_to_excel)
                        
                        ws = wb.create_sheet(name_str[2] ,-1)


                        channel_list = ["CH1", "CH2", "CH3"]

                        chart_anchors = ["B2","L2","I17"]
                        img_anchors = ["V3","V18","V33"]
                        chart_names = ["бип1 (CH1)","бип2 (CH2)", "бип1 - бип2"]
                        path_to_img  = os.path.join(resources, "screen").replace('\\','/')
                        
                        for cn in range(3):
                            path_to_dataframe = os.path.join(resources, str(cn+1)).replace('\\','/') +".txt"
                            if not(os.path.isfile(path_to_dataframe)):
                                break
                            else:
                                data = []
                                with open(path_to_dataframe, 'r', newline='') as csvfile:
                                    spamreader = csv.reader(csvfile, delimiter='\t', quotechar='|')
                                    for row in spamreader:
                                        data.append(row[:-1])
                                num_rows = len(data)
                                num_cols = len(data[0])


                                avg_pointer =  cn*2 + 3*(cn)+1 # указатели не надо пересчитывать но пусть будет

                                table_pointer = (3+2)*3 + cn*2+ num_cols*cn



                                ws.cell(column = table_pointer + 1, row = 1, value = channel_list[cn])
                                for j in range(0+2, num_rows+2):
                                    ws.cell(column = table_pointer+1, row = j,  value = j-2)

                                    for i in range(0+2,num_cols+2):
                                        ws.cell(column = table_pointer + i, row = 1,  value = i-2)
                                        ws.cell(column = table_pointer + i, row = j,  value = float(data[j-2][i-2].replace(',','.')))


                                    ws.cell(column = avg_pointer+1, row = j,  value = '=AVERAGE(' + get_column_letter(table_pointer+8+2) + str(j) + ':' + get_column_letter(table_pointer+12+2) + str(j) + ')'  )
                                    ws.cell(column = avg_pointer+2, row = j,  value = '=AVERAGE(' + get_column_letter(table_pointer+15+2) + str(j) + ':' + get_column_letter(table_pointer+25+2) + str(j) + ')'  )
                                    ws.cell(column = avg_pointer+3, row = j,  value = '=AVERAGE(' + get_column_letter(table_pointer+30+2) + str(j) + ':' + get_column_letter(table_pointer+70+2) + str(j) + ')'  )

                                    ws.cell(column = avg_pointer, row = j,  value =  j-2)


                                ws.cell(column = avg_pointer+1, row = 1, value = 'альфа 8-12')
                                ws.cell(column = avg_pointer+2, row = 1, value = 'бета 15-25')
                                ws.cell(column = avg_pointer+3, row = 1, value = 'гамма 30-70')


                                
                                
                                img = Image(path_to_img + str(cn+1) + '.png')
                                img.anchor = img_anchors[cn]
                                ws.add_image(img)


                                
                                c = LineChart()
                                c.title = chart_names[cn]


                                c.x_axis.delete = False
                                c.y_axis.delete = False


                                data = Reference(ws, min_col = avg_pointer+1, max_col = avg_pointer+3, min_row = 1, max_row = j) # адреса
                                c.add_data(data, titles_from_data = True)


                                s1 = c.series[0]

                                s1.graphicalProperties.solidFill = "FF0000"
                                s2 = c.series[1]
                                s2.graphicalProperties.solidFill = "0000FF"
                                s3 = c.series[2]
                                s3.graphicalProperties.solidFill = "00FF00"

                                ws.add_chart(c, anchor = chart_anchors[cn])
                                
                                ws.cell(column = 1, row = 1, value = name_str[4])
                            


                        
                        wb.save(path_to_excel)
                        try:
                            os.remove(resources + "/1.txt")
                        except OSError:
                            pass
                        try:
                            os.remove(resources + "/2.txt")
                        except OSError:
                            pass
                        try:
                            os.remove(resources + "/3.txt")
                        except OSError:
                            pass

        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
            wb.save(path_to_excel)

        


         
print('finish')
